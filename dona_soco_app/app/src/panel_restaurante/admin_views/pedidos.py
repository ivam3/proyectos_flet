import flet as ft
from database import obtener_pedidos, obtener_total_pedidos, actualizar_estado_pedido, actualizar_pago_pedido, obtener_datos_exportacion, obtener_menu
from components.notifier import init_pubsub, play_notification_sound, show_notification # Importar herramientas de notificación
import math
import csv
import datetime
import io
import os
import openpyxl
from fpdf import FPDF

def pedidos_view(page: ft.Page, export_file_picker: ft.FilePicker):
    """
    Vista del panel de administración para gestionar los pedidos con paginación y filtro.
    """
    rows_per_page = 12 
    current_page = 1
    total_pages = 1
    
    search_filter = ft.TextField(
        hint_text="Buscar por Cliente o Código",
        prefix_icon=ft.Icons.SEARCH,
        border_radius=20, height=40,
        text_size=14, content_padding=10, filled=True,
        text_style=ft.TextStyle(color=ft.Colors.BLACK),
        hint_style=ft.TextStyle(color=ft.Colors.GREY_600)
    )
    
    # --- LÓGICA HÍBRIDA DE EXPORTACIÓN ---
    # 1. Web/Escritorio: Usamos FilePicker (Descarga estándar/Diálogo SO)
    # 2. Android: Usamos Escritura Directa (FilePicker crashea en APK 0.80.0)

    # Configuración para Web/Escritorio
    file_picker = export_file_picker
    
    def on_file_picker_result(e):
        if e.path:
             show_notification(page, f"Archivo guardado exitosamente.", ft.Colors.GREEN)
        else:
             show_notification(page, "Operación cancelada.", ft.Colors.GREY_700)

    file_picker.on_result = on_file_picker_result

    # Dialogos para Android
    error_dialog = ft.AlertDialog(title=ft.Text("Error"), content=ft.Text(""))
    success_dialog = ft.AlertDialog(
        title=ft.Text("Descarga Exitosa", color=ft.Colors.GREEN), 
        content=ft.Text(""),
        actions=[ft.TextButton("Aceptar", on_click=lambda e: cerrar_dialogos())]
    )
    page.overlay.extend([error_dialog, success_dialog])

    # Diálogo de éxito de impresión
    print_success_dialog = ft.AlertDialog(
        title=ft.Text("Impresión Enviada", color=ft.Colors.GREEN, weight="bold"),
        content=ft.Column([
            ft.Icon(ft.Icons.PRINT_ROUNDED, size=50, color=ft.Colors.GREEN),
            ft.Text("Los tickets han sido enviados correctamente a las impresoras correspondientes:", text_align="center", color=ft.Colors.BLACK),
            ft.Text("", weight="bold", color=ft.Colors.BLUE_GREY_700, text_align="center")
        ], tight=True, horizontal_alignment="center"),
        actions=[ft.TextButton("Entendido", on_click=lambda e: setattr(print_success_dialog, "open", False) or page.update())]
    )
    page.overlay.append(print_success_dialog)

    def cerrar_dialogos():
        error_dialog.open = False
        success_dialog.open = False
        page.update()

    def mostrar_error(msg):
        error_dialog.content.value = str(msg)
        error_dialog.open = True
        page.update()

    def mostrar_exito_android(ruta):
        success_dialog.content.value = f"El archivo se ha guardado correctamente en:\n\n{ruta}"
        success_dialog.open = True
        page.update()

    def guardar_archivo_android(filename, content_bytes):
        rutas_a_probar = []
        # 1. Ruta estándar de Descargas
        rutas_a_probar.append(os.path.join("/storage/emulated/0/Download", filename))
        # 2. Ruta interna (Fallback)
        rutas_a_probar.append(os.path.join(os.getcwd(), filename))

        guardado = False
        ruta_final = ""

        for ruta in rutas_a_probar:
            try:
                print(f"DEBUG: Intentando guardar en {ruta}")
                os.makedirs(os.path.dirname(ruta), exist_ok=True)
                with open(ruta, "wb") as f:
                    f.write(content_bytes)
                guardado = True
                ruta_final = ruta
                break
            except Exception as e:
                print(f"DEBUG: Fallo al escribir en {ruta}: {e}")
                continue

        if guardado:
            mostrar_exito_android(ruta_final)
        else:
            mostrar_error("No se pudo guardar el archivo. Verifique permisos de almacenamiento.")

    async def descargar_archivo_web(filename, content_bytes, mime_type="application/octet-stream"):
        """Método compatible con Flet 0.24.1 usando launch_url en la misma ventana."""
        import base64
        try:
            b64 = base64.b64encode(content_bytes).decode()
            url = f"data:{mime_type};base64,{b64}"
            # En esta versión el argumento es web_popup_window_name
            await page.launch_url(url, web_popup_window_name="_self")
            show_notification(page, f"Descarga iniciada: {filename}", ft.Colors.GREEN)
        except Exception as e:
            print(f"Error en descarga web: {e}")
            show_notification(page, "Error al procesar descarga.", ft.Colors.RED)

    async def iniciar_exportacion(extension="csv"):
        print(f"DEBUG: Iniciando exportación {extension}")
        show_notification(page, f"Generando datos {extension.upper()}...", ft.Colors.BLUE_GREY_700)

        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            search_term = search_filter.value.strip() if search_filter.value else None
            
            print("DEBUG: Consultando base de datos...")
            datos = obtener_datos_exportacion(search_term=search_term)
            
            headers = [
                "Orden ID", "Código", "Fecha", "Cliente", "Teléfono", "Dirección", "Referencias", 
                "Estado", "Método Pago", "Paga Con", "Total Orden", "Motivo Cancelación",
                "Producto", "Cantidad", "Precio Unitario", "Subtotal Producto"
            ]
            
            content_bytes = None
            file_ext = extension
            mime = "text/csv"

            if extension == "csv":
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(headers)
                for row in datos:
                    writer.writerow(list(row.values()))
                content_bytes = output.getvalue().encode('utf-8')
                output.close()
                file_ext = "csv"
                mime = "text/csv"
            elif extension == "xlsm":
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Detalle de Ordenes"
                ws.append(headers)
                for row in datos:
                    ws.append(list(row.values()))
                output = io.BytesIO()
                wb.save(output)
                content_bytes = output.getvalue()
                output.close()
                file_ext = "xlsx" 
                mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
            if content_bytes:
                filename = f"reporte_{timestamp}.{file_ext}"
                
                # --- SELECTOR DE ESTRATEGIA ---
                plat = str(page.platform).lower() if page.platform else ""
                es_web = page.web
                
                # 1. Escritura Directa para Escritorio Nativo
                if plat in ["windows", "macos", "linux"] and not es_web:
                    try:
                        ruta_descargas = os.path.join(os.path.expanduser("~"), "Downloads", filename)
                        with open(ruta_descargas, "wb") as f:
                            f.write(content_bytes)
                        show_notification(page, f"Reporte guardado en Descargas.", ft.Colors.GREEN)
                        return
                    except Exception as e:
                        print(f"DEBUG: Fallo guardado directo en PC: {e}")

                # 2. Descarga Directa para Web (Solución al bloqueo de FilePicker)
                if es_web:
                    print("DEBUG: Usando descarga Base64 (Web).")
                    await descargar_archivo_web(filename, content_bytes, mime)
                elif plat in ["windows", "macos", "linux"]:
                    # Fallback FilePicker para escritorio nativo si falló el directo
                    await file_picker.save_file(
                        file_name=filename,
                        allowed_extensions=[file_ext],
                        src_bytes=content_bytes
                    )
                else:
                    print("DEBUG: Modo Móvil Nativo detectado. Usando Escritura Directa.")
                    guardar_archivo_android(filename, content_bytes)
            else:
                raise Exception("No se encontraron datos")

        except Exception as ex:
            print(f"DEBUG ERROR: {ex}")
            mostrar_error(f"Error: {ex}")

    # --- LÓGICA DE IMPRESIÓN (COCINA, FOODTRUCK, CAJA) ---
    async def imprimir_pedido(pedido):
        show_notification(page, "Iniciando impresión masiva...", ft.Colors.BLUE)

        try:
            # 1. Obtener mapa de productos para saber destino
            # Nota: Esto trae todo el menú. En producción optimizar con caché o mapa estático.
            menu_items = obtener_menu(solo_activos=False)
            # Map: "Nombre Producto" -> "cocina" | "foodtruck"
            # Normalizamos nombres para evitar fallos por espacios
            product_map = { m['nombre'].strip().lower(): m.get('printer_target', 'cocina') for m in menu_items }
            
            # 2. Clasificar items del pedido
            # pedido['detalles'] es la lista de objetos dict con keys: producto, cantidad, precio_unitario
            items_caja = []     # Todos
            items_cocina = []   # Target cocina
            items_foodtruck = [] # Target foodtruck
            
            for item in pedido.get('detalles', []):
                items_caja.append(item)
                
                # Extraer nombre base si tiene extras (ej: "Tacos (Sin cebolla)")
                nombre_full = item['producto']
                nombre_base = nombre_full.split("(")[0].strip().lower()
                
                target = product_map.get(nombre_base, 'cocina') # Default a cocina si no se encuentra
                
                if target == 'foodtruck':
                    items_foodtruck.append(item)
                else:
                    items_cocina.append(item)

            # 3. Función Simulata de Envío a Impresora (Logica ESC/POS iría aquí)
            def enviar_a_impresora(nombre_impresora, items, es_ticket_completo=False):
                if not items: return
                
                print(f"--- IMPRIMIENDO EN {nombre_impresora.upper()} ---")
                print(f"Pedido #{pedido['id']} - {pedido['nombre_cliente']}")
                for i in items:
                    print(f"- {i['cantidad']} x {i['producto']}")
                print("------------------------------------------")
                
                # AQUI IRÍA LA CONEXIÓN TCP/IP o USB CON LA IMPRESORA
                # Ejemplo pseudocodigo:
                # printer = NetworkPrinter("192.168.1.xxx")
                # printer.text(f"Pedido #{pedido['id']}\n")
                # ...
            
            # 4. Ejecutar impresiones
            # CAJA (Siempre imprime todo)
            enviar_a_impresora("CAJA", items_caja, es_ticket_completo=True)
            
            # COCINA (Solo si hay items)
            if items_cocina:
                enviar_a_impresora("COCINA", items_cocina)
            
            # FOODTRUCK (Solo si hay items)
            if items_foodtruck:
                enviar_a_impresora("FOODTRUCK", items_foodtruck)

            # Mostrar confirmación visual clara en pantalla
            areas = ["Caja"]
            if items_cocina: areas.append("Cocina")
            if items_foodtruck: areas.append("Foodtruck")
            
            # Actualizar texto del diálogo (segundo control de la columna)
            print_success_dialog.content.controls[2].value = " + ".join(areas)
            print_success_dialog.open = True
            page.update()

        except Exception as ex:
            print(f"ERROR IMPRESION: {ex}")
            show_notification(page, f"Error imprimiendo: {ex}", ft.Colors.RED)

    def print_handler(pedido):
        async def handler(e):
            await imprimir_pedido(pedido)
        return handler

    async def generar_y_guardar_pdf(pedido):
        show_notification(page, "Generando PDF...", ft.Colors.BLUE_GREY_700)

        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("helvetica", size=12)
            pdf.set_margins(10, 10, 10)
            
            pdf.set_font("helvetica", 'B', 16)
            pdf.cell(0, 10, text=f"Detalle del Pedido #{pedido['id']}", align='C', new_x="LMARGIN", new_y="NEXT")
            pdf.ln(5)
            
            pdf.set_font("helvetica", size=12)
            pdf.cell(0, 10, text=f"Fecha: {pedido['fecha']}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 10, text=f"Código: {pedido['codigo_seguimiento']}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(5)
            
            pdf.set_font("helvetica", 'B', 14)
            pdf.cell(0, 10, text="Datos del Cliente", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("helvetica", size=12)
            pdf.cell(0, 8, text=f"Nombre: {pedido['nombre_cliente']}", new_x="LMARGIN", new_y="NEXT")
            
            direccion = (pedido['direccion'] or "").encode('latin-1', 'replace').decode('latin-1')
            pdf.multi_cell(w=pdf.epw, h=8, text=f"Dirección: {direccion}")
            pdf.ln(5)
            
            pdf.set_font("helvetica", 'B', 14)
            pdf.cell(0, 10, text="Productos", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("helvetica", size=11)
            
            detalles = pedido["detalles_productos"].split(" | ") if pedido["detalles_productos"] else []
            for item in detalles:
                safe_item = item.encode('latin-1', 'replace').decode('latin-1')
                pdf.multi_cell(w=pdf.epw, h=7, text=f"- {safe_item}", border=0, align='L')
                pdf.ln(1)
            
            pdf.ln(5)
            pdf.set_font("helvetica", 'B', 16)
            pdf.cell(0, 10, text=f"Total: ${pedido['total']:.2f}", align='R', new_x="LMARGIN", new_y="NEXT")

            pdf_bytes = pdf.output()
            filename = f"pedido_{pedido['id']}.pdf"
            
            # --- SELECTOR DE ESTRATEGIA ---
            plat = str(page.platform).lower() if page.platform else ""
            es_web = page.web

            # 1. Escritura Directa para Escritorio Nativo
            if plat in ["windows", "macos", "linux"] and not es_web:
                try:
                    ruta_descargas = os.path.join(os.path.expanduser("~"), "Downloads", filename)
                    with open(ruta_descargas, "wb") as f:
                        f.write(pdf_bytes)
                    show_notification(page, f"PDF guardado en Descargas.", ft.Colors.GREEN)
                    return
                except Exception as e:
                    print(f"DEBUG: Fallo guardado directo PDF en PC: {e}")

            # 2. Descarga Directa para Web
            if es_web:
                print("DEBUG: Usando descarga Base64 (PDF).")
                # Usamos application/octet-stream para forzar descarga y evitar visor PDF negro
                await descargar_archivo_web(filename, pdf_bytes, "application/octet-stream")
            elif plat in ["windows", "macos", "linux"]:
                # Fallback FilePicker para nativo
                await file_picker.save_file(
                    file_name=filename,
                    allowed_extensions=["pdf"],
                    src_bytes=pdf_bytes
                )
            else:
                print("DEBUG: Modo Móvil Nativo detectado (PDF). Usando Escritura Directa.")
                guardar_archivo_android(filename, pdf_bytes)
            
        except Exception as ex:
             mostrar_error(f"Error PDF: {ex}")

    def create_pdf_handler(pedido):
        async def handler(e):
            await generar_y_guardar_pdf(pedido)
        return handler

    async def export_csv_click(e):
        await iniciar_exportacion("csv")

    async def export_xlsm_click(e):
        await iniciar_exportacion("xlsm")

    pedidos_data_table = ft.DataTable(
        heading_row_color=ft.Colors.ORANGE_100,
        heading_row_height=60,
        columns=[
            ft.DataColumn(ft.Text("ID", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Código", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Cliente", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Fecha", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Total", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Pago", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Estado", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Acciones", weight="bold", color=ft.Colors.BLACK)),
        ],
        rows=[],
        border=ft.Border.all(1, ft.Colors.GREY_300),
        vertical_lines=ft.border.BorderSide(1, ft.Colors.GREY_300),
        horizontal_lines=ft.border.BorderSide(1, ft.Colors.GREY_300),
    )

    def close_details_dialog(e):
        details_dialog.open = False
        page.update()

    details_dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text("Detalles del Pedido", color=ft.Colors.BLACK),
        content=ft.Column(), 
        actions=[ft.TextButton("Cerrar", on_click=close_details_dialog, style=ft.ButtonStyle(color=ft.Colors.BROWN_700))],
        actions_alignment=ft.MainAxisAlignment.END,
    )
    page.overlay.append(details_dialog)

    def open_details_dialog(e, pedido):
        detalles_productos_lista = pedido["detalles_productos"].split(" | ") if pedido["detalles_productos"] else []
        metodo = pedido["metodo_pago"] or "N/A"
        paga_con = pedido["paga_con"] or 0
        cambio = paga_con - pedido["total"] if metodo == "efectivo" and paga_con else 0
        
        pago_info = [ft.Text(f"Método de Pago: {metodo.capitalize()}", weight="bold", color=ft.Colors.BLACK)]
        if metodo == "efectivo":
            pago_info.append(ft.Text(f"Paga con: ${paga_con:.2f}", color=ft.Colors.BLACK))
            pago_info.append(ft.Text(f"Cambio: ${cambio:.2f}", color=ft.Colors.GREEN_700 if cambio >= 0 else ft.Colors.RED))

        # Motivo cancelación
        motivo_widget = ft.Container()
        if str(pedido['estado']).lower() == "cancelado" and pedido.get('motivo_cancelacion'):
            motivo_widget = ft.Column([
                ft.Divider(),
                ft.Text("❌ PEDIDO CANCELADO", color=ft.Colors.RED, weight="bold"),
                ft.Text(f"Motivo: {pedido.get('motivo_cancelacion', '')}", color=ft.Colors.RED_700)
            ])

        details_dialog.title = ft.Text(f"Pedido #{pedido['id']}", color=ft.Colors.BLACK)
        details_dialog.content = ft.Container(
            content=ft.Column([
                ft.Text(f"Código Seguimiento: {pedido['codigo_seguimiento']}", weight="bold", size=16, color=ft.Colors.BLUE_GREY_700),
                ft.Divider(),
                ft.Text(f"Cliente: {pedido['nombre_cliente']}", weight="bold", color=ft.Colors.BLACK),
                ft.Text(f"Teléfono: {pedido['telefono']}", color=ft.Colors.BLACK),
                ft.Text(f"Dirección: {pedido['direccion']}", color=ft.Colors.BLACK),
                ft.Divider(),
                ft.Text("Productos:", weight="bold", color=ft.Colors.BLACK),
                ft.Column([ft.Text(f"- {item}", color=ft.Colors.BLACK) for item in detalles_productos_lista]),
                ft.Divider(),
                ft.Column(pago_info),
                ft.Divider(),
                ft.Text(f"Total: ${pedido['total']:.2f}", weight="bold", size=16, color=ft.Colors.BLACK),
                ft.Text(f"Estado Actual: {pedido['estado']}", weight="bold", color=ft.Colors.BLACK),
                motivo_widget
            ], scroll="auto"),
            height=400, width=350,
        )
        details_dialog.open = True
        page.update()

    def close_confirmation_dialog(e=None):
        confirmation_dialog.open = False
        page.update()

    def confirm_status_change(e, pedido_id, new_status, motivo=None):
        actualizar_estado_pedido(pedido_id, new_status, motivo)
        close_confirmation_dialog()
        cargar_pedidos()
        show_notification(page, f"Estado actualizado a {new_status}", ft.Colors.GREEN)

    confirmation_dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text("Cambiar Estado", color=ft.Colors.BLACK),
        content=ft.Container(),
        actions=[],
    )
    page.overlay.append(confirmation_dialog)

    def open_status_dialog(e, pedido):
        def set_status(status):
            return lambda e: confirm_status_change(e, pedido['id'], status)

        def show_cancel_reason_dialog(e):
            reason_field = ft.TextField(label="Motivo de cancelación", multiline=True, autofocus=True, text_style=ft.TextStyle(color=ft.Colors.BLACK))
            
            def confirm_cancel(e):
                if not reason_field.value or not reason_field.value.strip():
                    reason_field.error_text = "Debes ingresar un motivo"
                    reason_field.update()
                    return
                
                # Cerrar primero el diálogo de motivo
                cancel_dialog.open = False
                page.update()
                
                # Luego procesar el cambio que cierra el diálogo principal
                confirm_status_change(None, pedido['id'], "Cancelado", reason_field.value.strip())

            cancel_dialog = ft.AlertDialog(
                title=ft.Text("Confirmar Cancelación", color=ft.Colors.BLACK),
                content=ft.Container(content=reason_field, height=150),
                actions=[
                    ft.TextButton("Volver", on_click=lambda e: setattr(cancel_dialog, "open", False) or page.update(), style=ft.ButtonStyle(color=ft.Colors.BROWN_700)),
                    ft.FilledButton("Confirmar", on_click=confirm_cancel, style=ft.ButtonStyle(bgcolor=ft.Colors.RED, color=ft.Colors.WHITE))
                ]
            )
            page.overlay.append(cancel_dialog)
            cancel_dialog.open = True
            page.update()

        confirmation_dialog.title = ft.Text(f"Cambiar estado pedido #{pedido['id']}", color=ft.Colors.BLACK)
        
        # Deshabilitar si ya está cancelado
        is_canceled = str(pedido['estado']).lower() == "cancelado"
        
        controls = [
            ft.FilledButton("Pendiente", on_click=set_status("Pendiente"), width=200, style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE), disabled=is_canceled),
            ft.FilledButton("Preparando", on_click=set_status("Preparando"), width=200, style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE), disabled=is_canceled),
            ft.FilledButton("En Camino", on_click=set_status("En Camino"), width=200, style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE), disabled=is_canceled),
            ft.FilledButton("Entregado", on_click=set_status("Entregado"), width=200, style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE), disabled=is_canceled),
            ft.Divider(),
            ft.FilledButton("Cancelar Pedido", on_click=show_cancel_reason_dialog, width=200, style=ft.ButtonStyle(bgcolor=ft.Colors.RED, color=ft.Colors.WHITE), disabled=is_canceled),
        ]

        if is_canceled:
            controls.append(ft.Text("Este pedido ya fue cancelado.", color=ft.Colors.RED, weight="bold"))

        confirmation_dialog.content = ft.Column(controls, height=350, alignment=ft.MainAxisAlignment.CENTER, scroll="auto")
        
        confirmation_dialog.actions = [ft.TextButton("Cerrar", on_click=close_confirmation_dialog)]
        confirmation_dialog.open = True
        page.update()

    txt_page_info = ft.Text("Página 1 de 1", color=ft.Colors.BLACK)
    
    def change_page(delta):
        nonlocal current_page
        new_page = current_page + delta
        if 1 <= new_page <= total_pages:
            current_page = new_page
            cargar_pedidos()

    btn_prev = ft.IconButton(icon=ft.Icons.ARROW_BACK, icon_color=ft.Colors.BLACK, on_click=lambda e: change_page(-1))
    btn_next = ft.IconButton(icon=ft.Icons.ARROW_FORWARD, icon_color=ft.Colors.BLACK, on_click=lambda e: change_page(1))

    def cargar_pedidos():
        nonlocal current_page, total_pages
        search_term = search_filter.value.strip() if search_filter.value else None
        
        total_items = obtener_total_pedidos(search_term=search_term)
        
        total_pages = math.ceil(total_items / rows_per_page) if total_items > 0 else 1
        current_page = max(1, min(current_page, total_pages))
        
        offset = (current_page - 1) * rows_per_page
        
        pedidos = obtener_pedidos(limit=rows_per_page, offset=offset, search_term=search_term)
        
        pedidos_data_table.rows.clear()
        for p in pedidos:
            es_cancelado = str(p['estado']).lower() == "cancelado"
            pedidos_data_table.rows.append(
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text(str(p['id']), color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Text(p['codigo_seguimiento'], color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Text(p['nombre_cliente'], color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Text(str(p['fecha']), color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Text(f"${p['total']:.2f}", color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Text(str(p['metodo_pago']).capitalize(), color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Text(p['estado'], color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Row([
                        ft.IconButton(ft.Icons.VISIBILITY, icon_color=ft.Colors.BLUE_GREY_700, on_click=lambda e, p=p: open_details_dialog(e, p)),
                        ft.IconButton(
                            ft.Icons.EDIT, 
                            icon_color=ft.Colors.GREY_400 if es_cancelado else ft.Colors.BLUE_GREY_700, 
                            disabled=es_cancelado, 
                            on_click=lambda e, p=p: open_status_dialog(e, p)
                        ),
                        ft.IconButton(ft.Icons.PRINT, icon_color=ft.Colors.BLUE, tooltip="Imprimir Tickets (Cocina/Foodtruck)", on_click=print_handler(p)),
                        ft.IconButton(ft.Icons.PICTURE_AS_PDF, icon_color=ft.Colors.RED_700, on_click=create_pdf_handler(p))
                    ])),
                ])
            )
        txt_page_info.value = f"Página {current_page} de {total_pages}"
        btn_prev.disabled = current_page <= 1
        btn_next.disabled = current_page >= total_pages
        page.update()

    cargar_pedidos()

    # --- SUBSCRIPCIÓN A NOTIFICACIONES ---
    async def on_new_order(message):
        if message == "nuevo_pedido":
            # Reproducir sonido
            play_notification_sound(page)
            # Mostrar alerta visual
            show_notification(page, "🔔 ¡Nuevo Pedido Recibido!", ft.Colors.GREEN_700)
            
            # Recargar tabla
            cargar_pedidos()
            page.update()

            # Imprimir automaticamente
            try:
                # Obtener el ultimo pedido para imprimirlo
                # Asumimos que obtener_pedidos devuelve ordenados por fecha descendente
                pedidos_recientes = obtener_pedidos(limit=1)
                if pedidos_recientes:
                    ultimo_pedido = pedidos_recientes[0]
                    # Llamar a la funcion de impresion existente
                    await imprimir_pedido(ultimo_pedido)
            except Exception as e:
                print(f"Error en impresion automatica: {e}")

    pubsub = init_pubsub(page)
    pubsub.subscribe(on_new_order)

    content_container = ft.Container(
        padding=20,
        border=ft.Border.all(1, ft.Colors.OUTLINE_VARIANT),
        border_radius=15,
        expand=True,
        content=ft.Column(
            controls=[
                ft.Text("Gestión de pedidos", size=20, weight="bold", color=ft.Colors.BLUE_GREY_900),
                search_filter,
                ft.Row([
                    ft.FilledButton("Filtrar", on_click=lambda e: cargar_pedidos(), style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE)),
                    ft.FilledButton("Limpiar", on_click=lambda e: (setattr(search_filter, "value", ""), cargar_pedidos()), style=ft.ButtonStyle(bgcolor=ft.Colors.RED, color=ft.Colors.WHITE)),
                    ft.IconButton(ft.Icons.REFRESH, on_click=lambda e: cargar_pedidos(), icon_color=ft.Colors.BLUE_GREY_700, tooltip="Actualizar lista"),
                ], spacing=10),
                # Botones de exportación
                ft.Row([
                    ft.FilledButton("CSV", icon=ft.Icons.DOWNLOAD, on_click=export_csv_click, expand=True, style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE)),
                    ft.FilledButton("Excel", icon=ft.Icons.TABLE_VIEW, on_click=export_xlsm_click, expand=True, style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE))
                ], spacing=10),
                # Área de la tabla
                ft.Column(
                    [
                        ft.Row([pedidos_data_table], scroll=ft.ScrollMode.ALWAYS)
                    ]
                ),
                ft.Row([btn_prev, txt_page_info, btn_next], alignment=ft.MainAxisAlignment.CENTER),
            ],
            scroll="auto",
            expand=True,
            spacing=15
        )
    )

    return ft.Column([content_container], expand=True)