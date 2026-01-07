import flet as ft
from database import obtener_pedidos, obtener_total_pedidos, actualizar_estado_pedido, actualizar_pago_pedido, obtener_datos_exportacion
import math
import csv
import datetime
import io
import os
import openpyxl
from fpdf import FPDF

def pedidos_view(page: ft.Page, export_file_picker: ft.FilePicker):
    """
    Vista del panel de administración para gestionar los pedidos con paginación y filtro.
    """
    rows_per_page = 12 
    current_page = 1
    total_pages = 1
    
    search_filter = ft.TextField(
        hint_text="Buscar por Cliente o Código",
        prefix_icon=ft.Icons.SEARCH,
        border_radius=20, height=40,
        text_size=14, content_padding=10, filled=True,
        text_style=ft.TextStyle(color=ft.Colors.BLACK),
        hint_style=ft.TextStyle(color=ft.Colors.GREY_600)
    )
    
    # Usamos el FilePicker pasado por parámetro (el que ya está en el overlay de la página)
    file_picker = export_file_picker

    def on_file_picker_result(e):
        if e.path:
             page.snack_bar = ft.SnackBar(ft.Text(f"Archivo guardado: {e.path}", color=ft.Colors.WHITE), bgcolor=ft.Colors.GREEN)
        else:
             page.snack_bar = ft.SnackBar(ft.Text("Operación finalizada.", color=ft.Colors.WHITE))
        
        page.snack_bar.open = True
        page.update()

    # Sobrescribimos el evento para esta vista
    file_picker.on_result = on_file_picker_result

    async def iniciar_exportacion(extension="csv"):
        page.snack_bar = ft.SnackBar(ft.Text(f"Generando reporte {extension.upper()}...", color=ft.Colors.WHITE))
        page.snack_bar.open = True
        page.update()

        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            search_term = search_filter.value.strip() if search_filter.value else None
            datos = obtener_datos_exportacion(search_term=search_term)
            
            headers = [
                "Orden ID", "Código", "Fecha", "Cliente", "Teléfono", "Dirección", "Referencias", 
                "Estado", "Método Pago", "Paga Con", "Total Orden", "Motivo Cancelación",
                "Producto", "Cantidad", "Precio Unitario", "Subtotal Producto"
            ]
            
            content_bytes = None
            file_ext = extension

            if extension == "csv":
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(headers)
                for row in datos:
                    writer.writerow(list(row))
                content_bytes = output.getvalue().encode('utf-8')
                output.close()
                file_ext = "csv"
            elif extension == "xlsm":
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Detalle de Ordenes"
                ws.append(headers)
                for row in datos:
                    ws.append(list(row))
                output = io.BytesIO()
                wb.save(output)
                content_bytes = output.getvalue()
                output.close()
                file_ext = "xlsx" 
            
            if content_bytes:
                # En Flet 0.80.x para Android, a veces el parámetro es 'data' o 'src_bytes'
                # Intentamos con src_bytes que es el estándar moderno para Web/Mobile
                await file_picker.save_file(
                    dialog_title=f"Guardar {file_ext.upper()}",
                    file_name=f"reporte_{timestamp}.{file_ext}",
                    allowed_extensions=[file_ext],
                    src_bytes=content_bytes
                )
            else:
                raise Exception("Sin datos")

        except Exception as ex:
            page.snack_bar = ft.SnackBar(ft.Text(f"Error: {ex}", color=ft.Colors.WHITE), bgcolor=ft.Colors.RED)
            page.snack_bar.open = True
            page.update()

    async def generar_y_guardar_pdf(pedido):
        page.snack_bar = ft.SnackBar(ft.Text("Generando PDF...", color=ft.Colors.WHITE))
        page.snack_bar.open = True
        page.update()

        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("helvetica", size=12)
            pdf.set_margins(10, 10, 10)
            
            pdf.set_font("helvetica", 'B', 16)
            pdf.cell(0, 10, text=f"Detalle del Pedido #{pedido['id']}", align='C', new_x="LMARGIN", new_y="NEXT")
            pdf.ln(5)
            
            pdf.set_font("helvetica", size=12)
            pdf.cell(0, 10, text=f"Fecha: {pedido['fecha']}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 10, text=f"Código: {pedido['codigo_seguimiento']}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(5)
            
            pdf.set_font("helvetica", 'B', 14)
            pdf.cell(0, 10, text="Datos del Cliente", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("helvetica", size=12)
            pdf.cell(0, 8, text=f"Nombre: {pedido['nombre_cliente']}", new_x="LMARGIN", new_y="NEXT")
            
            direccion = (pedido['direccion'] or "").encode('latin-1', 'replace').decode('latin-1')
            pdf.multi_cell(w=pdf.epw, h=8, text=f"Dirección: {direccion}")
            pdf.ln(5)
            
            pdf.set_font("helvetica", 'B', 14)
            pdf.cell(0, 10, text="Productos", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("helvetica", size=11) # Reducir un poco el tamaño para descripciones largas
            
            detalles = pedido["detalles_productos"].split(" | ") if pedido["detalles_productos"] else []
            for item in detalles:
                # Limpiar y preparar el texto
                safe_item = item.encode('latin-1', 'replace').decode('latin-1')
                # multi_cell con w=pdf.epw forzará el salto de línea automático dentro de los márgenes
                pdf.multi_cell(w=pdf.epw, h=7, text=f"- {safe_item}", border=0, align='L')
                pdf.ln(1) # Pequeño espacio entre productos
            
            pdf.ln(5)
            pdf.set_font("helvetica", 'B', 16)
            pdf.cell(0, 10, text=f"Total: ${pedido['total']:.2f}", align='R', new_x="LMARGIN", new_y="NEXT")

            pdf_bytes = bytes(pdf.output())
            
            await file_picker.save_file(
                dialog_title=f"Guardar PDF #{pedido['id']}",
                file_name=f"pedido_{pedido['id']}.pdf",
                allowed_extensions=["pdf"],
                src_bytes=pdf_bytes
            )
            
        except Exception as ex:
             page.snack_bar = ft.SnackBar(ft.Text(f"Error PDF: {ex}", color=ft.Colors.WHITE), bgcolor=ft.Colors.RED)
             page.snack_bar.open = True
             page.update()

    def create_pdf_handler(pedido):
        async def handler(e):
            await generar_y_guardar_pdf(pedido)
        return handler

    async def export_csv_click(e):
        await iniciar_exportacion("csv")

    async def export_xlsm_click(e):
        await iniciar_exportacion("xlsm")

    pedidos_data_table = ft.DataTable(
        heading_row_color=ft.Colors.ORANGE_100,
        heading_row_height=60,
        columns=[
            ft.DataColumn(ft.Text("ID", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Código", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Cliente", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Fecha", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Total", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Pago", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Estado", weight="bold", color=ft.Colors.BLACK)),
            ft.DataColumn(ft.Text("Acciones", weight="bold", color=ft.Colors.BLACK)),
        ],
        rows=[],
        border=ft.Border.all(1, ft.Colors.GREY_300),
        vertical_lines=ft.border.BorderSide(1, ft.Colors.GREY_300),
        horizontal_lines=ft.border.BorderSide(1, ft.Colors.GREY_300),
    )

    def close_details_dialog(e):
        details_dialog.open = False
        page.update()

    details_dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text("Detalles del Pedido", color=ft.Colors.BLACK),
        content=ft.Column(), 
        actions=[ft.TextButton("Cerrar", on_click=close_details_dialog, style=ft.ButtonStyle(color=ft.Colors.BROWN_700))],
        actions_alignment=ft.MainAxisAlignment.END,
    )
    page.overlay.append(details_dialog)

    def open_details_dialog(e, pedido):
        detalles_productos_lista = pedido["detalles_productos"].split(" | ") if pedido["detalles_productos"] else []
        metodo = pedido["metodo_pago"] or "N/A"
        paga_con = pedido["paga_con"] or 0
        cambio = paga_con - pedido["total"] if metodo == "efectivo" and paga_con else 0
        
        pago_info = [ft.Text(f"Método de Pago: {metodo.capitalize()}", weight="bold", color=ft.Colors.BLACK)]
        if metodo == "efectivo":
            pago_info.append(ft.Text(f"Paga con: ${paga_con:.2f}", color=ft.Colors.BLACK))
            pago_info.append(ft.Text(f"Cambio: ${cambio:.2f}", color=ft.Colors.GREEN_700 if cambio >= 0 else ft.Colors.RED))

        details_dialog.title = ft.Text(f"Pedido #{pedido['id']}", color=ft.Colors.BLACK)
        details_dialog.content = ft.Container(
            content=ft.Column([
                ft.Text(f"Código Seguimiento: {pedido['codigo_seguimiento']}", weight="bold", size=16, color=ft.Colors.BLUE_GREY_700),
                ft.Divider(),
                ft.Text(f"Cliente: {pedido['nombre_cliente']}", weight="bold", color=ft.Colors.BLACK),
                ft.Text(f"Teléfono: {pedido['telefono']}", color=ft.Colors.BLACK),
                ft.Text(f"Dirección: {pedido['direccion']}", color=ft.Colors.BLACK),
                ft.Divider(),
                ft.Text("Productos:", weight="bold", color=ft.Colors.BLACK),
                ft.Column([ft.Text(f"- {item}", color=ft.Colors.BLACK) for item in detalles_productos_lista]),
                ft.Divider(),
                ft.Column(pago_info),
                ft.Divider(),
                ft.Text(f"Total: ${pedido['total']:.2f}", weight="bold", size=16, color=ft.Colors.BLACK),
                ft.Text(f"Estado Actual: {pedido['estado']}", weight="bold", color=ft.Colors.BLACK),
            ], scroll="auto"),
            height=400, width=350,
        )
        details_dialog.open = True
        page.update()

    def close_confirmation_dialog(e=None):
        confirmation_dialog.open = False
        page.update()

    def confirm_status_change(e, pedido_id, new_status, motivo=None):
        actualizar_estado_pedido(pedido_id, new_status, motivo)
        close_confirmation_dialog()
        cargar_pedidos()
        page.snack_bar = ft.SnackBar(ft.Text(f"Estado actualizado a {new_status}", color=ft.Colors.WHITE))
        page.snack_bar.open = True
        page.update()

    confirmation_dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text("Cambiar Estado", color=ft.Colors.BLACK),
        content=ft.Container(),
        actions=[],
    )
    page.overlay.append(confirmation_dialog)

    def open_status_dialog(e, pedido):
        def set_status(status):
            return lambda e: confirm_status_change(e, pedido['id'], status)

        confirmation_dialog.title = ft.Text(f"Cambiar estado pedido #{pedido['id']}", color=ft.Colors.BLACK)
        confirmation_dialog.content = ft.Column([
            ft.FilledButton("Pendiente", on_click=set_status("Pendiente"), width=200, style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE)),
            ft.FilledButton("Preparando", on_click=set_status("Preparando"), width=200, style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE)),
            ft.FilledButton("En Camino", on_click=set_status("En Camino"), width=200, style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE)),
            ft.FilledButton("Entregado", on_click=set_status("Entregado"), width=200, style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE)),
        ], height=250, alignment=ft.MainAxisAlignment.CENTER)
        
        confirmation_dialog.actions = [ft.TextButton("Cerrar", on_click=close_confirmation_dialog)]
        confirmation_dialog.open = True
        page.update()

    txt_page_info = ft.Text("Página 1 de 1", color=ft.Colors.BLACK)
    
    def change_page(delta):
        nonlocal current_page
        new_page = current_page + delta
        if 1 <= new_page <= total_pages:
            current_page = new_page
            cargar_pedidos()

    btn_prev = ft.IconButton(icon=ft.Icons.ARROW_BACK, icon_color=ft.Colors.BLACK, on_click=lambda e: change_page(-1))
    btn_next = ft.IconButton(icon=ft.Icons.ARROW_FORWARD, icon_color=ft.Colors.BLACK, on_click=lambda e: change_page(1))

    def cargar_pedidos():
        nonlocal current_page, total_pages
        search_term = search_filter.value.strip() if search_filter.value else None
        
        total_items = obtener_total_pedidos(search_term=search_term)
        
        total_pages = math.ceil(total_items / rows_per_page) if total_items > 0 else 1
        current_page = max(1, min(current_page, total_pages))
        
        offset = (current_page - 1) * rows_per_page
        
        pedidos = obtener_pedidos(limit=rows_per_page, offset=offset, search_term=search_term)
        
        pedidos_data_table.rows.clear()
        for p in pedidos:
            pedidos_data_table.rows.append(
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text(str(p['id']), color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Text(p['codigo_seguimiento'], color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Text(p['nombre_cliente'], color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Text(str(p['fecha']), color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Text(f"${p['total']:.2f}", color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Text(str(p['metodo_pago']).capitalize(), color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Text(p['estado'], color=ft.Colors.BLACK)),
                    ft.DataCell(ft.Row([
                        ft.IconButton(ft.Icons.VISIBILITY, icon_color=ft.Colors.BLUE_GREY_700, on_click=lambda e, p=p: open_details_dialog(e, p)),
                        ft.IconButton(ft.Icons.EDIT, icon_color=ft.Colors.BLUE_GREY_700, on_click=lambda e, p=p: open_status_dialog(e, p)),
                        ft.IconButton(ft.Icons.PICTURE_AS_PDF, icon_color=ft.Colors.RED_700, on_click=create_pdf_handler(p))
                    ])),
                ])
            )
        txt_page_info.value = f"Página {current_page} de {total_pages}"
        btn_prev.disabled = current_page <= 1
        btn_next.disabled = current_page >= total_pages
        page.update()

    cargar_pedidos()

    content_container = ft.Container(
        padding=20,
        border=ft.Border.all(1, ft.Colors.OUTLINE_VARIANT),
        border_radius=15,
        expand=True,
        content=ft.Column(
            controls=[
                ft.Text("Gestión de pedidos", size=20, weight="bold", color=ft.Colors.BLUE_GREY_900),
                search_filter,
                ft.Row([
                    ft.FilledButton("Filtrar", on_click=lambda e: cargar_pedidos(), style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE)),
                    ft.FilledButton("Limpiar", on_click=lambda e: (setattr(search_filter, "value", ""), cargar_pedidos()), style=ft.ButtonStyle(bgcolor=ft.Colors.RED, color=ft.Colors.WHITE)),
                    ft.IconButton(ft.Icons.REFRESH, on_click=lambda e: cargar_pedidos(), icon_color=ft.Colors.BLUE_GREY_700, tooltip="Actualizar lista"),
                ], spacing=10),
                ft.Row([
                    ft.FilledButton("CSV", icon=ft.Icons.DOWNLOAD, on_click=export_csv_click, expand=True, style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE)),
                    ft.FilledButton("Excel", icon=ft.Icons.TABLE_VIEW, on_click=export_xlsm_click, expand=True, style=ft.ButtonStyle(bgcolor=ft.Colors.BROWN_700, color=ft.Colors.WHITE))
                ], spacing=10),
                ft.Column(
                    [
                        ft.Row([pedidos_data_table], scroll=ft.ScrollMode.ALWAYS)
                    ],
                    scroll=ft.ScrollMode.AUTO,
                    expand=True
                ),
                ft.Row([btn_prev, txt_page_info, btn_next], alignment=ft.MainAxisAlignment.CENTER),
            ],
            scroll="auto",
            expand=True,
            spacing=15
        )
    )

    return ft.Column([content_container], expand=True)
